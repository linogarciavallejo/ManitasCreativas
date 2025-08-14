#!/usr/bin/env python3
"""
Script to extract registration-fee payment dates and amounts from an
Excel workbook and produce INSERT SQL statements.  The workbook
``Pagos-Parvulos2-Kinder-A.xlsx`` contains student names in cells
``B3`` through ``B21`` and their corresponding school registration-fee
amounts in cells ``I3`` through ``I21``.  The cell comments attached
to the ``I`` column hold the date on which the payment was received.

This script performs the following steps for each student:

1. Normalize the student's name (remove accents, punctuation and
   collapse multiple spaces) to enable a robust match against the
   ``NombreCompleto`` column in the provided ``data-*.csv`` file.
2. Look up the student's ``Id`` in the CSV based on the normalized
   name.  The CSV file must contain two columns: ``Id`` and
   ``NombreCompleto``.
    3. Read the registration-fee cell value from a configurable
       column (``I`` by default) and convert it into a floating
       point amount.  Values may include a currency symbol
       (e.g. ``Q250,00``) and use a comma as the decimal
       separator.
    4. Extract the payment date from the fee cell's comment.  Dates are
   expected in the format ``DD MONTH YYYY`` (e.g. ``13 NOVIEMBRE 2024``)
   or in ``DD/MM/YYYY`` form.  If a date of the former form is
   present it is preferred; otherwise a ``DD/MM/YYYY`` date is used.
   The extracted date is converted into ISO ``YYYY-MM-DD`` format for
   insertion into the SQL statement.
    5. Generate an ``INSERT`` statement for the student using the
   provided template from ``insert-pago-example.sql``.  Only the
   ``Fecha``, ``Monto`` and ``AlumnoId`` fields are modified; all
   other fields retain the original values from the example.  If no
   date is found in the comment the student is skipped.

Usage::

    python generate_sql.py --excel <path-to-xlsx> --csv <path-to-csv> \
        [--sql-template <template.sql>] [--rubro-id <id>] \
        [--month <month>] [--year <year>] \
        [--cell-range <start:end>] [--fee-col <column>]

The script writes the resulting SQL statements to a ``.sql`` file whose
name is derived from the Excel filename.  Use ``--cell-range`` to
specify the range of name cells to process (including the column
letter) and ``--fee-col`` to override the column used for fee
amounts and comments.  The default behaviour uses the same offset
between name and amount columns as the sample template (``B`` →
``I``).  Fallback month and year values may be provided to
synthetically generate dates when comments are missing.
"""

import argparse
import unicodedata
import re
from typing import Dict, Optional

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def remove_accents(text: str) -> str:
    """Remove accents and diacritics from a Unicode string.

    Normalizes the string to NFKD form and drops combining characters
    to produce an ASCII‐only representation.  Useful for matching
    names irrespective of accents.

    Parameters
    ----------
    text: str
        Input string potentially containing accented characters.

    Returns
    -------
    str
        The input string without any accents.
    """
    nfkd_form = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd_form if not unicodedata.combining(c))


def normalize_name(name: str) -> str:
    """Normalize a student's name for comparison.

    This function uppercases the name, removes accents and
    punctuation (commas and periods) and collapses multiple
    whitespace characters into single spaces.

    Parameters
    ----------
    name: str
        The raw name string to normalize.

    Returns
    -------
    str
        A normalized representation of the name.
    """
    # Strip leading/trailing whitespace and uppercase
    s = name.strip().upper()
    # Remove accents
    s = remove_accents(s)
    # Remove punctuation (commas and periods)
    s = re.sub(r'[.,]', '', s)
    # Collapse multiple whitespace characters into one space
    s = re.sub(r'\s+', ' ', s)
    return s


def build_name_to_id_map(csv_path: str) -> Dict[str, int]:
    """Build a mapping from normalized student name to ID from the CSV.

    The CSV is expected to contain columns ``Id`` and ``NombreCompleto``.

    Parameters
    ----------
    csv_path: str
        Path to the CSV file containing student IDs and names.

    Returns
    -------
    Dict[str, int]
        A dictionary mapping normalized names to their corresponding ID.
    """
    df = pd.read_csv(csv_path)
    name_to_id = {}
    for _, row in df.iterrows():
        raw_name = str(row['NombreCompleto'])
        norm_name = normalize_name(raw_name)
        name_to_id[norm_name] = int(row['Id'])
    return name_to_id


# Mapping of Spanish month names to their numeric equivalent
SPANISH_MONTH_NAMES = {
    1: 'ENERO',
    2: 'FEBRERO',
    3: 'MARZO',
    4: 'ABRIL',
    5: 'MAYO',
    6: 'JUNIO',
    7: 'JULIO',
    8: 'AGOSTO',
    9: 'SEPTIEMBRE',
    10: 'OCTUBRE',
    11: 'NOVIEMBRE',
    12: 'DICIEMBRE',
}
# Reverse mapping: month name (no accents, uppercase) -> number
SPANISH_MONTH_NUMBERS = {remove_accents(name): num for num, name in SPANISH_MONTH_NAMES.items()}

# Regular expressions to detect dates in comments
PATTERN_DAY_MONTH_YEAR = re.compile(
    r'(?P<day>[0-3]?\d)\s+(?P<month>[A-ZÁÉÍÓÚÑ]+)\s+(?P<year>\d{4})', re.IGNORECASE
)
PATTERN_NUMERIC_DATE = re.compile(
    r'(?P<day>[0-3]?\d)[/-](?P<month>[01]?\d)[/-](?P<year>\d{4})'
)


def parse_date_from_comment(comment_text: str) -> Optional[str]:
    """Extract a payment date from a cell comment.

    The function looks for dates in two formats: ``DD MONTH YYYY``
    (where ``MONTH`` is a Spanish month name) and ``DD/MM/YYYY`` or
    ``DD-MM-YYYY``.  If a date of the first form is found, it is
    preferred; otherwise a slash‐delimited date is used.  The returned
    value is an ISO date string (``YYYY-MM-DD``).

    Parameters
    ----------
    comment_text: str
        The raw comment text attached to a cell.

    Returns
    -------
    Optional[str]
        ISO formatted date string (YYYY-MM-DD) if a date is found;
        otherwise ``None``.
    """
    if not comment_text:
        return None
    # Normalize comment: remove accents and uppercase for matching
    normalized = remove_accents(comment_text).upper()
    # Try to match the "DD MONTH YYYY" pattern first
    m = PATTERN_DAY_MONTH_YEAR.search(normalized)
    if m:
        day = int(m.group('day'))
        month_name = m.group('month')
        year = int(m.group('year'))
        month_num = SPANISH_MONTH_NUMBERS.get(month_name)
        if month_num is None:
            # If the month name isn't recognized just return the raw string
            return f"{day:02d} {month_name} {year}"
        return f"{year}-{month_num:02d}-{day:02d}"
    # Fallback: look for numeric dates with slash or dash
    m2 = PATTERN_NUMERIC_DATE.search(normalized)
    if m2:
        day = int(m2.group('day'))
        month = int(m2.group('month'))
        year = int(m2.group('year'))
        return f"{year}-{month:02d}-{day:02d}"
    return None


def parse_amount(value) -> Optional[float]:
    """Convert a cell value into a float amount.

    Values may be integers, floats or strings containing a currency
    symbol and a comma decimal separator (e.g. ``Q250,00``).  Any
    thousand separators (periods) are stripped and comma decimal
    separators are converted to dots.  If the value cannot be parsed
    returns ``None``.

    Parameters
    ----------
    value: Any
        The raw cell value from the workbook.

    Returns
    -------
    Optional[float]
        A floating point representation of the amount or ``None`` if
        conversion fails.
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    # Remove any non‐digit, non comma, non dot, non minus characters
    s_clean = re.sub(r'[^0-9,.-]', '', s)
    # Remove thousand separators (periods) and convert comma to dot for decimals
    # e.g. "Q1.234,56" -> "1234.56"
    s_clean = s_clean.replace('.', '')
    s_clean = s_clean.replace(',', '.')
    if not s_clean:
        return None
    try:
        return float(s_clean)
    except ValueError:
        return None


def load_template(sql_path: str) -> str:
    """Load the sample SQL insert statement.

    The template file must contain exactly one complete ``INSERT``
    statement.  Comments and formatting are preserved.  The caller
    should replace the placeholders for the ``Fecha``, ``Monto`` and
    ``AlumnoId`` fields when generating statements for each student.

    Parameters
    ----------
    sql_path: str
        Path to the SQL file containing the example INSERT statement.

    Returns
    -------
    str
        The raw contents of the SQL file.
    """
    with open(sql_path, 'r', encoding='utf-8') as f:
        contents = f.read()
    # Many SQL templates include header comments (e.g. a SELECT statement
    # or notes about what the statement does).  When generating multiple
    # INSERT statements these comments quickly become redundant.  We
    # therefore trim any leading comment lines (beginning with ``--``)
    # until we encounter the actual ``INSERT INTO`` line.  The
    # remainder of the file is returned unchanged, preserving
    # indentation and inline comments.
    lines = contents.splitlines(True)
    start_idx = 0
    for idx, line in enumerate(lines):
        # Look for the first line that begins with "INSERT INTO"
        if line.strip().upper().startswith('INSERT INTO'):
            start_idx = idx
            break
    trimmed = ''.join(lines[start_idx:])
    return trimmed


def generate_statement(
    template: str,
    fecha: str,
    monto: float,
    alumno_id: int,
    rubro_id: int,
) -> str:
    """Produce a customized INSERT statement for one student.

    Starting from the provided template, this function substitutes
    values for several fields:

    - ``Fecha``: replaces the first ``CURRENT_TIMESTAMP`` occurrence.
    - ``Monto``: replaces the first numeric literal with a decimal
      component (assumed to be the amount in the template).
    - ``AlumnoId``: replaces the numeric literal on the ``AlumnoId`` line.
    - ``MedioPago``: forces the value to ``1``.
    - ``Notas``: sets an empty string.
    - ``RubroId``: parameterised via ``rubro_id``.
    - ``EsColegiatura``: forced to ``false``.
    - ``MesColegiatura`` and ``AnioColegiatura``: both replaced with
      ``NULL``.

    Other fields are left untouched.  Regular expressions are used to
    perform targeted replacements while preserving indentation and
    inline comments.

    Parameters
    ----------
    template: str
        The raw SQL template.
    fecha: str
        ISO formatted date (YYYY-MM-DD) to substitute for the ``Fecha`` field.
    monto: float
        The amount to substitute for the ``Monto`` field.
    alumno_id: int
        Student ID to substitute for the ``AlumnoId`` field.
    rubro_id: int
        RubroId value to use in the statement.

    Returns
    -------
    str
        A fully formatted INSERT statement customised for the given student.
    """
    out = template
    # 1. Replace the first occurrence of CURRENT_TIMESTAMP with the supplied date (Fecha)
    out = out.replace('CURRENT_TIMESTAMP', f"'{fecha}'", 1)
    # 2. Replace the first floating point number (the amount placeholder) with our amount
    out = re.sub(
        r'\b\d+\.\d+\b',
        lambda m: f"{monto:.2f}" if m.group(0) else m.group(0),
        out,
        count=1,
    )
    # 3. Replace AlumnoId value
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- AlumnoId',
        lambda m: f"\n{m.group(1)}{alumno_id},                              -- AlumnoId",
        out,
        count=1,
    )
    # 4. Force MedioPago to 1
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- MedioPago',
        lambda m: f"\n{m.group(1)}1,                              -- MedioPago",
        out,
        count=1,
    )
    # 5. Set Notas to empty string
    out = re.sub(
        r"'[^']*',\s*-- Notas",
        "'',              -- Notas",
        out,
        count=1,
    )
    # 6. Replace RubroId with provided rubro_id
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- RubroId',
        lambda m: f"\n{m.group(1)}{rubro_id},                              -- RubroId",
        out,
        count=1,
    )
    # 7. Ensure EsColegiatura is false
    out = re.sub(
        r'\n(\s*)true,\s*-- EsColegiatura',
        lambda m: f"\n{m.group(1)}false,                           -- EsColegiatura",
        out,
        count=1,
    )
    # 8. Set MesColegiatura to NULL
    out = re.sub(
        r'EXTRACT\(MONTH FROM CURRENT_DATE\),\s*-- MesColegiatura',
        'NULL,                              -- MesColegiatura',
        out,
        count=1,
    )
    # 9. Set AnioColegiatura to NULL
    out = re.sub(
        r'EXTRACT\(YEAR FROM CURRENT_DATE\),\s*-- AnioColegiatura',
        'NULL,                              -- AnioColegiatura',
        out,
        count=1,
    )
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate INSERT statements from Excel and CSV data.')
    parser.add_argument('--excel', required=True, help='Path to the Excel workbook (xlsx)')
    parser.add_argument('--csv', required=True, help='Path to the CSV file containing student IDs and names')
    parser.add_argument('--sql-template', default='insert-pago-example.sql', help='Path to the SQL template file')
    parser.add_argument('--rubro-id', type=int, default=8, help='RubroId to use in the generated statements')
    parser.add_argument('--month', type=int, help='Fallback month (1-12) to use when a payment date is missing')
    parser.add_argument('--year', type=int, help='Fallback year to use when a payment date is missing')
    parser.add_argument(
        '--cell-range',
        default='B3:B21',
        help=(
            'Cell range for the student names in "<col><row>:<col><row>" format '
            '(e.g. "B3:B21").  The amount/comment column is derived from '
            'the same relative offset used in the example template (B -> I).'
        ),
    )
    parser.add_argument(
        '--fee-col',
        default=None,
        help=(
            'Column letter to read the registration fee and its comment from. '
            'If omitted, the fee column is derived from the name column using the '
            'offset in the template (e.g. B->I).'
        ),
    )
    args = parser.parse_args()

    # Build name -> ID mapping
    name_to_id = build_name_to_id_map(args.csv)

    # Load workbook
    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active

    # Load template SQL
    template = load_template(args.sql_template)

    # Parse the cell range argument (e.g. "B3:B21").  Extract the column letter
    # and start/end row numbers.  Use the first cell's column for names and
    # compute the amount column by applying the offset between the template's
    # name and amount columns (B -> I).  Accept either colon or dash as
    # separator.
    cell_range = args.cell_range
    try:
        if ':' in cell_range:
            first_cell, last_cell = cell_range.split(':', 1)
        elif '-' in cell_range:
            first_cell, last_cell = cell_range.split('-', 1)
        else:
            first_cell = last_cell = cell_range
        # Extract column letters and row numbers from the cell references
        import re
        cell_re = re.compile(r'(?i)([A-Z]+)(\d+)')
        m_start = cell_re.fullmatch(first_cell.strip())
        m_end = cell_re.fullmatch(last_cell.strip())
        if not (m_start and m_end):
            raise ValueError
        name_col_letter = m_start.group(1).upper()
        start_row = int(m_start.group(2))
        # We allow the end column to be omitted or different; rows from second cell
        end_row = int(m_end.group(2))
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        # Determine the fee (amount) column letter.  If the user supplied
        # --fee-col, use that directly; otherwise derive it from the
        # name column by applying the same offset used in the original
        # template (B -> I corresponds to an offset of 7 columns).
        if args.fee_col:
            amount_col_letter = args.fee_col.strip().upper()
        else:
            name_col_index = column_index_from_string(name_col_letter)
            amount_col_index = name_col_index + (column_index_from_string('I') - column_index_from_string('B'))
            amount_col_letter = get_column_letter(amount_col_index)
    except Exception:
        print(
            f"Invalid cell range '{cell_range}'. Expected format '<col><row>:<col><row>' e.g. 'B3:B21'."
        )
        return

    # Determine output filename based on the Excel filename
    # Use the stem (remove extension) and append .sql
    import os
    excel_basename = os.path.basename(args.excel)
    stem, _ = os.path.splitext(excel_basename)
    output_filename = f"{stem}.sql"

    statements = []  # collect all generated statements
    # Iterate over the rows of interest based on the provided range
    for row in range(start_row, end_row + 1):
        # Build cell references dynamically based on the parsed range
        name_cell_ref = f"{name_col_letter}{row}"
        amount_cell_ref = f"{amount_col_letter}{row}"
        name_cell = ws[name_cell_ref].value
        amount_cell = ws[amount_cell_ref].value
        # Skip rows without a student name
        if not name_cell:
            continue
        # Normalize name to look up ID
        norm_name = normalize_name(str(name_cell))
        alumno_id = name_to_id.get(norm_name)
        if alumno_id is None:
            # Skip silently if no ID found
            continue
        # Parse amount
        monto = parse_amount(amount_cell)
        if monto is None:
            # Skip silently if amount cannot be parsed
            continue
        # Extract date from comment
        comment = ws[amount_cell_ref].comment
        fecha = parse_date_from_comment(comment.text if comment else '')
        # If no date found, use fallback month/year if provided
        if fecha is None:
            if args.month and args.year:
                # Use the first day of the specified month/year
                fecha = f"{args.year}-{args.month:02d}-01"
            else:
                # If no fallback specified, skip this row
                continue
        # Generate statement
        stmt = generate_statement(
            template,
            fecha,
            monto,
            alumno_id,
            args.rubro_id,
        )
        statements.append(stmt.strip())

    # Write all statements to the output file
    if statements:
        with open(output_filename, 'w', encoding='utf-8') as f:
            f.write('\n\n'.join(statements))
        print(f"Generated {len(statements)} statements and wrote them to {output_filename}")
    else:
        print("No statements were generated.")


if __name__ == '__main__':
    main()