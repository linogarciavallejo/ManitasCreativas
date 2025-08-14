#!/usr/bin/env python3
"""
Extended script to extract multiple registration-fee payments per student
from an Excel workbook and produce ``INSERT`` SQL statements.  This
version supports processing an arbitrary number of consecutive fee
columns.  Each fee column's heading (assumed to be in row 2) is
interpreted as a Spanish month name (e.g. ``Abril``, ``Mayo``,
``Junio``).  The ``MesColegiatura`` value in the generated statement
is derived from this month name: ``Abril`` → ``4``, ``Mayo`` → ``5``,
etc.  The ``AnioColegiatura`` field is set to ``2025`` and
``EsColegiatura`` is forced to ``true`` for all statements.

The workflow:

1. Normalize student names and match them against IDs in the CSV.
2. For each row in the name range, iterate over the specified fee
   columns.  Each non-empty fee cell is treated as a separate payment.
3. Parse the amount from the cell and extract a date from the
   comment (falling back to the first of the month/year if needed).
   The entire comment text is preserved (with whitespace normalised)
   and written into the ``Notas`` field of the payment record.
4. Generate an ``INSERT`` statement for each payment using a template,
   substituting ``Fecha``, ``Monto``, ``AlumnoId``, ``RubroId``,
   ``MesColegiatura`` (derived from the column header), ``AnioColegiatura``
   (2025), ``EsColegiatura`` (true) and ``Notas`` (full comment).

Usage::

    python generate_sql_multi_mes.py --excel <xlsx> --csv <csv> \
        [--sql-template <file.sql>] [--rubro-id <id>] \
        [--month <month>] [--year <year>] \
        [--cell-range <start:end>] [--fee-cols <start:end>]

``--cell-range`` defines the range of name cells (e.g. ``B3:B21``).
``--fee-cols`` defines the start and end columns (e.g. ``J-L``).  All
columns within the range will be processed.  For each column, the
script reads its header text (in row 2) and maps the Spanish month
name to a numeric ``MesColegiatura`` (e.g. ``Abril`` → ``4``).  If
``--fee-cols`` is omitted, the first fee column is derived from the
name column using the sample template offset (B→I) and only that
single column is processed.
"""

import argparse
import unicodedata
import re
from typing import Dict, Optional, List

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def remove_accents(text: str) -> str:
    nfkd_form = unicodedata.normalize('NFKD', text)
    return ''.join(c for c in nfkd_form if not unicodedata.combining(c))


def normalize_name(name: str) -> str:
    s = name.strip().upper()
    s = remove_accents(s)
    s = re.sub(r'[.,]', '', s)
    s = re.sub(r'\s+', ' ', s)
    return s


def build_name_to_id_map(csv_path: str) -> Dict[str, int]:
    df = pd.read_csv(csv_path)
    name_to_id: Dict[str, int] = {}
    for _, row in df.iterrows():
        norm_name = normalize_name(str(row['NombreCompleto']))
        name_to_id[norm_name] = int(row['Id'])
    return name_to_id


SPANISH_MONTH_NAMES = {
    1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL', 5: 'MAYO',
    6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO', 9: 'SEPTIEMBRE',
    10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
}
SPANISH_MONTH_NUMBERS = {remove_accents(name): num for num, name in SPANISH_MONTH_NAMES.items()}

PATTERN_DAY_MONTH_YEAR = re.compile(
    r'(?P<day>[0-3]?\d)\s+(?P<month>[A-ZÁÉÍÓÚÑ]+)\s+(?P<year>\d{4})', re.IGNORECASE
)
PATTERN_NUMERIC_DATE = re.compile(
    r'(?P<day>[0-3]?\d)[/-](?P<month>[01]?\d)[/-](?P<year>\d{4})'
)


def parse_date_from_comment(comment_text: str) -> Optional[str]:
    if not comment_text:
        return None
    normalized = remove_accents(comment_text).upper()
    m = PATTERN_DAY_MONTH_YEAR.search(normalized)
    if m:
        day = int(m.group('day'))
        month_name = m.group('month')
        year = int(m.group('year'))
        month_num = SPANISH_MONTH_NUMBERS.get(month_name)
        if month_num is None:
            return f"{day:02d} {month_name} {year}"
        return f"{year}-{month_num:02d}-{day:02d}"
    m2 = PATTERN_NUMERIC_DATE.search(normalized)
    if m2:
        day = int(m2.group('day'))
        month = int(m2.group('month'))
        year = int(m2.group('year'))
        return f"{year}-{month:02d}-{day:02d}"
    return None


def parse_amount(value) -> Optional[float]:
    """Parse a cell value into a monetary amount.

    This function returns ``None`` for values that should not be
    considered monetary amounts.  By default, amounts are required to
    be numeric.  Strings containing non-numeric tokens (e.g. ``XX``,
    ``BECA``, ``PENDIENTE``), or date-like values (containing ``/``)
    are treated as invalid and will result in ``None``.  Negative and
    zero amounts are also rejected.

    Parameters
    ----------
    value : Any
        The raw cell value from the workbook.

    Returns
    -------
    Optional[float]
        The parsed monetary amount, or ``None`` if the value is
        invalid.
    """
    if value is None:
        return None
    # Allow numeric types directly but reject non-positive values
    if isinstance(value, (int, float)):
        amt = float(value)
        return amt if amt > 0 else None
    s = str(value).strip()
    # Reject strings that contain a slash (likely dates) or known non-numeric tokens
    if not s:
        return None
    s_lower = s.lower()
    if '/' in s_lower:
        return None
    if s_lower in {'xx', 'x', 'beca', 'pendiente', 'na', 'n/a', '-'}:
        return None
    # Remove all characters except digits, comma, period and minus
    s_clean = re.sub(r'[^0-9,.-]', '', s)
    # Normalize thousand separators and decimal comma
    # If both comma and period appear, assume comma is decimal if it appears after the last period
    if ',' in s_clean and '.' in s_clean:
        # Determine which is decimal separator by position of last comma vs last period
        if s_clean.rfind(',') > s_clean.rfind('.'):
            # Remove thousands separators (periods) and convert decimal comma to dot
            s_clean = s_clean.replace('.', '')
            s_clean = s_clean.replace(',', '.')
        else:
            # Remove thousands separators (commas)
            s_clean = s_clean.replace(',', '')
    else:
        # Only one of comma or period is present
        if s_clean.count(',') and not re.search(r',\d{1,2}$', s_clean):
            # comma used as thousands separator
            s_clean = s_clean.replace(',', '')
        # replace remaining comma with dot as decimal separator
        s_clean = s_clean.replace(',', '.')
    # After cleaning, ensure string still contains digits
    if not re.search(r'\d', s_clean):
        return None
    try:
        amt = float(s_clean)
        return amt if amt > 0 else None
    except ValueError:
        return None


def load_template(sql_path: str) -> str:
    with open(sql_path, 'r', encoding='utf-8') as f:
        contents = f.read()
    lines = contents.splitlines(True)
    start_idx = 0
    for idx, line in enumerate(lines):
        if line.strip().upper().startswith('INSERT INTO'):
            start_idx = idx
            break
    return ''.join(lines[start_idx:])


def generate_statement(
    template: str,
    fecha: str,
    monto: float,
    alumno_id: int,
    rubro_id: int,
    mes_colegiatura: int,
    anio_colegiatura: int,
    es_colegiatura: bool,
    notas: str,
) -> str:
    """Customise the template with values for one payment."""
    out = template
    # Replace CURRENT_TIMESTAMP (Fecha)
    out = out.replace('CURRENT_TIMESTAMP', f"'{fecha}'", 1)
    # Replace first floating point number (Monto)
    out = re.sub(
        r'\b\d+\.\d+\b',
        lambda m: f"{monto:.2f}" if m.group(0) else m.group(0),
        out,
        count=1,
    )
    # Replace AlumnoId
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- AlumnoId',
        lambda m: f"\n{m.group(1)}{alumno_id},                              -- AlumnoId",
        out,
        count=1,
    )
    # Force MedioPago to 1
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- MedioPago',
        lambda m: f"\n{m.group(1)}1,                              -- MedioPago",
        out,
        count=1,
    )
    # Replace Notas with provided comment text (escaping single quotes)
    notas_escaped = notas.replace("'", "''") if notas is not None else ''
    out = re.sub(
        r"'[^']*',\s*-- Notas",
        lambda m: f"'{notas_escaped}',              -- Notas",
        out,
        count=1,
    )
    # Replace RubroId
    out = re.sub(
        r'\n(\s*)(\d+),\s*-- RubroId',
        lambda m: f"\n{m.group(1)}{rubro_id},                              -- RubroId",
        out,
        count=1,
    )
    # Set EsColegiatura
    out = re.sub(
        r'\n(\s*)(true|false),\s*-- EsColegiatura',
        lambda m: f"\n{m.group(1)}{'true' if es_colegiatura else 'false'},                           -- EsColegiatura",
        out,
        count=1,
    )
    # Set MesColegiatura
    out = re.sub(
        r'\n(\s*)([^,\n]+),\s*-- MesColegiatura',
        lambda m: f"\n{m.group(1)}{mes_colegiatura},                              -- MesColegiatura",
        out,
        count=1,
    )
    # Set AnioColegiatura
    out = re.sub(
        r'\n(\s*)([^,\n]+),\s*-- AnioColegiatura',
        lambda m: f"\n{m.group(1)}{anio_colegiatura},                              -- AnioColegiatura",
        out,
        count=1,
    )
    return out


def parse_cell_range(cell_range: str):
    if ':' in cell_range:
        first_cell, last_cell = cell_range.split(':', 1)
    elif '-' in cell_range:
        first_cell, last_cell = cell_range.split('-', 1)
    else:
        first_cell = last_cell = cell_range
    cell_re = re.compile(r'(?i)([A-Z]+)(\d+)')
    m_start = cell_re.fullmatch(first_cell.strip())
    m_end = cell_re.fullmatch(last_cell.strip())
    if not (m_start and m_end):
        raise ValueError
    name_col_letter = m_start.group(1).upper()
    start_row = int(m_start.group(2))
    end_row = int(m_end.group(2))
    if start_row > end_row:
        start_row, end_row = end_row, start_row
    return name_col_letter, start_row, end_row


def parse_fee_cols(fee_cols: Optional[str], name_col_letter: str) -> List[str]:
    if fee_cols:
        if ':' in fee_cols:
            start, end = fee_cols.split(':', 1)
        elif '-' in fee_cols:
            start, end = fee_cols.split('-', 1)
        else:
            start = end = fee_cols
        start = start.strip().upper()
        end = end.strip().upper()
        start_idx = column_index_from_string(start)
        end_idx = column_index_from_string(end)
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
        return [get_column_letter(i) for i in range(start_idx, end_idx + 1)]
    # Derive single column using B→I offset
    name_idx = column_index_from_string(name_col_letter)
    fee_idx = name_idx + (column_index_from_string('I') - column_index_from_string('B'))
    return [get_column_letter(fee_idx)]


def main() -> None:
    parser = argparse.ArgumentParser(description='Generate INSERT statements from Excel and CSV data (multi fee columns with MesColegiatura).')
    parser.add_argument('--excel', required=True, help='Path to the Excel workbook (xlsx)')
    parser.add_argument('--csv', required=True, help='Path to the CSV file with student IDs and names')
    parser.add_argument('--sql-template', default='insert-pago-example.sql', help='Path to the SQL template file')
    parser.add_argument('--rubro-id', type=int, default=8, help='RubroId for the generated statements')
    parser.add_argument('--month', type=int, help='Fallback month (1-12) for dates missing in comments')
    parser.add_argument('--year', type=int, help='Fallback year for dates missing in comments')
    parser.add_argument('--cell-range', default='B3:B21', help='Range of name cells (e.g. B3:B21)')
    parser.add_argument('--fee-cols', default=None, help='Range of fee columns (e.g. J-L) or single column (J)')
    args = parser.parse_args()

    name_to_id = build_name_to_id_map(args.csv)
    wb = openpyxl.load_workbook(args.excel)
    ws = wb.active
    template = load_template(args.sql_template)
    try:
        name_col_letter, start_row, end_row = parse_cell_range(args.cell_range)
    except Exception:
        print(f"Invalid cell range '{args.cell_range}'.")
        return
    try:
        fee_cols_list = parse_fee_cols(args.fee_cols, name_col_letter)
    except Exception:
        print(f"Invalid fee column specification '{args.fee_cols}'.")
        return

    import os
    excel_basename = os.path.basename(args.excel)
    stem, _ = os.path.splitext(excel_basename)
    # Determine output filename; if a file with the base name exists, append
    # a counter suffix (e.g. "filename-2.sql", "filename-3.sql", etc.)
    base_output = f"{stem}.sql"
    output_filename = base_output
    if os.path.exists(output_filename):
        counter = 2
        while True:
            candidate = f"{stem}-{counter}.sql"
            if not os.path.exists(candidate):
                output_filename = candidate
                break
            counter += 1

    statements: List[str] = []
    # row 2 contains the month headers for fee columns
    header_row = 2
    for row in range(start_row, end_row + 1):
        name_val = ws[f"{name_col_letter}{row}"].value
        if not name_val:
            continue
        norm_name = normalize_name(str(name_val))
        alumno_id = name_to_id.get(norm_name)
        if alumno_id is None:
            continue
        for fee_col_letter in fee_cols_list:
            # Determine month number for this fee column by reading its header (row 2)
            header_val = ws[f"{fee_col_letter}{header_row}"].value
            if not header_val:
                # no header: cannot determine month
                continue
            header_norm = remove_accents(str(header_val)).upper().strip()
            # consider only the first word in case of extra text
            first_word = header_norm.split()[0] if header_norm else ''
            mes_colegiatura = SPANISH_MONTH_NUMBERS.get(first_word)
            if mes_colegiatura is None:
                # Attempt to parse numeric month from header
                m = re.match(r'.*?(\d{1,2}).*', header_norm)
                if m:
                    mes_colegiatura = int(m.group(1))
                else:
                    # Unknown month name: skip this fee column
                    continue
            anio_colegiatura = 2025
            es_colegiatura = True
            fee_value = ws[f"{fee_col_letter}{row}"].value
            monto = parse_amount(fee_value)
            if monto is None:
                continue
            comment_obj = ws[f"{fee_col_letter}{row}"].comment
            comment_text = comment_obj.text if comment_obj else ''
            # Extract the date from the full comment
            fecha = parse_date_from_comment(comment_text)
            # Derive Notas: keep full comment but normalise whitespace (replace newlines with spaces)
            notas = ' '.join(comment_text.split()) if comment_text else ''
            if fecha is None:
                if args.month and args.year:
                    fecha = f"{args.year}-{args.month:02d}-01"
                else:
                    continue
            stmt = generate_statement(
                template,
                fecha,
                monto,
                alumno_id,
                args.rubro_id,
                mes_colegiatura,
                anio_colegiatura,
                es_colegiatura,
                notas,
            )
            statements.append(stmt.strip())

    if statements:
        with open(output_filename, 'w', encoding='utf-8') as f:
            f.write('\n\n'.join(statements))
        print(f"Generated {len(statements)} statements and wrote them to {output_filename}")
    else:
        print('No statements were generated.')


if __name__ == '__main__':
    main()